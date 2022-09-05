"""
Curso: Introdução à programação Python

Projeto Python 2
Escreva um programa em Python que crie planilhas dentro do arquivo orcamento.xls
que se encontra na pasta planilhas que voce criou. As planilhas devem ter os
nomes a seguir: -receitas, -despesas, -resultado.

Autor: Guilherme Medeiros Lionço
Data: 05/09/2022
Versão:0.0.1

Observação. O meu módulo openpyxl não consegue trabalhar com arquivos .xls,
somente funciona com arquivos .xlsx
"""

print ("\nEste programa abrirá uma planilha e irá manusear as abas")

#importando o módulo openpyxl para trabalhar com planilhas
from openpyxl import load_workbook

#abrindo a planilha
wb=load_workbook("orcamento.xlsx")

#renomeando a aba sheet 1
ws1 = wb.active
ws1.title = 'receitas'

#criando as outras 2 abas
ws2=wb.create_sheet("despesas")
ws3=wb.create_sheet("resultado")

#salvando o arquivo
wb.save("orcamento.xlsx")

print ("\nA criação de abas na planilha foi finalizada.")

